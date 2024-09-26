from datetime import datetime

import openpyxl
import os
from colorist import red, effect_underline, Color, effect_bold

def filter_file(filename):
    #filters file names
    if not filename.startswith("~") and filename.endswith("xlsx"):
        return True
    return False

current_season = "fall"
current_semester = "2024-2025"

while True:
    # which module 269 or 110
    target_module = int(input("Enter Module number: "))
    modules = [269, 110]
    if target_module not in modules:
        print(f"Please choose appropriate module between the specified choices: {modules}")
        continue
    # getting week number to record attendance for the corresponding week
    target_week = int(input("Enter week number: "))
    week_range = range(1, 15)
    if target_week not in week_range:
        print(f"Please choose appropriate range between the specified choice: {week_range}")
        continue
    # dictionary where the attendance files at
    directory = f"D:\\astro\\Astro\\uni tutoring\\M{target_module}\\attendance\\{current_semester}\\{current_season}"
    print(f"Fetch from directory: {directory}")
    break
# using this list for displaying purposes once the students is recorded as attended
# it is also used to write the data to a temp file in case of corruption
# it stores a tuple(studentId, sectionNumber, StudentName)
recorded_students = []

while True:

    target_id = int(input("Enter student id: "))

    if target_id == -1:
        break

    # using this counter to indicate that the student id wasn't found in
    # the excell files if this counter value is 2 (failure_value) which is the number of excel files we are searching in
    failure_value = 2
    notFoundCounter = 0

    attendance_files = list(filter(filter_file, os.listdir(directory)))
    # looping through files within the directory and opening only xlsx files
    for filename in attendance_files:
        # getting different information about the file first
        # has an extra column so I am checking before finding cells and committing changes
        is269 = 269 == target_module
        # id starts within the 6th row
        id_row =  6
        # id found in column 2 in all files
        id_column = 2
        # 269 files has an extra column
        week_column = 5 + target_week if is269 else 4 + target_week

        name_column = id_column + 1
        section_number = filename[10:12]

        # opening the file and activating the current sheet
        try:
            wb = openpyxl.load_workbook(os.path.join(directory, filename))
            sheet = wb.active
        except PermissionError:
            red("Error while Opening files: Please make sure attendance files are closed while taking the "
                "attendance")
            quit(-1)


        # this indicator used to exit the loops once student's id is found and marked attended
        # it is useful to immediately save the changes once the id found to avoid redundancy
        changeMade = False

        # looping through each record within the table
        # starting from the rows and columns where student ids are found
        # again to avoid redundancy
        for record in sheet.iter_rows(min_row=id_row, min_col=id_column, max_col=id_column):

            # looping through each cell
            # and comparing its value with the target student's id
            for cell in record:
                # once found we get the current row and column to move from there
                # to get more information like the name
                if cell.value == target_id:
                    row = cell.row
                    column = cell.column

                    # marking the student as attended for the specific week
                    sheet.cell(row=row, column=week_column, value="p")
                    # getting students name
                    name = sheet.cell(row, name_column).value
                    # displaying to console that the operation was successful with bold text
                    effect_bold(
                        f"Name: {name}\nSection: {section_number}\nhas been successfully marked as present for "
                        f"week {target_week}\n\n")
                    # here we set the changeMade to true, so we can immediately stop looping
                    # and ask for the next student's id
                    changeMade = True
                    # adding the record to list for the reasons set earlier
                    recorded_students.append((target_id, section_number, name))
                    break
            if changeMade:
                break
        if changeMade:
            # saving the file and resetting the counter, so it doesn't reach 4
            # indicating that student's id was found and changes has been made
            try:
                wb.save(os.path.join(directory, filename))
                notFoundCounter = 0
            except PermissionError:
                red(f"Error while Saving files: {name} ({target_id}) was not recorded as an attendee, please make sure attendance files are closed while taking the attendance")
            break
        else:
            # if the id is not found within a file then it will be incremented
            # until it reaches 4 indicating that the id is not within any file
            # then we counter
            notFoundCounter += 1

        # if id is not found within the 4 files
        if notFoundCounter == failure_value:
            red(f"Student with id {target_id} is not found")
            notFoundCounter = 0


def formatRecord(currentIndex):
    """Formatting information saved within the recorded students list"""
    return f"{currentIndex + 1}) {recorded_students[currentIndex][0]} - {recorded_students[currentIndex][1]} -" \
           f" {recorded_students[currentIndex][2]} "


def checkTempFileAvailability():
    """Creates the temp file within the directory if it doesn't exist and returns its path"""
    file_path = os.path.join(directory, "temp")
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    return file_path


def writeRecordToFile():
    """Saving information recorded in recorded_students list into a file in case of corruption & checking"""

    path = checkTempFileAvailability()

    # construct the file name without invalid characters
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    file_name = f"Week{target_week}_M{target_module}_{timestamp}.txt"

    with open(os.path.join(path, file_name), 'w') as f:
        for index in range(len(recorded_students)):
            f.write(formatRecord(index) + "\n")


# removing duplicates
recorded_students = list(set(recorded_students))

# displaying how many students where recorded and saving info into a file
if recorded_students:
    writeRecordToFile()
    effect_underline("Recorded students", Color.GREEN)
    for i in range(len(recorded_students)):
        record = formatRecord(i)
        effect_underline(record, Color.GREEN)
    effect_underline(f"Recorded {len(recorded_students)} students successfully", Color.GREEN)
